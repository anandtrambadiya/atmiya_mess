from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_login import login_required, current_user
from models import db, User, Pass, PassUsageLog, MealCount, OneTimeCollection, SystemSettings, DailyNote
from datetime import datetime, date, timedelta
from functools import wraps
from sqlalchemy import func
from zoneinfo import ZoneInfo
import io, calendar

IST = ZoneInfo("Asia/Kolkata")

def today_ist():
    return datetime.now(IST).date()

boss_bp = Blueprint('boss', __name__)

def boss_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('user_type') != 'admin' or current_user.role != 'boss':
            flash('Boss access required', 'error')
            return redirect(url_for('staff.dashboard'))
        return f(*args, **kwargs)
    return decorated

# ── Settings ──────────────────────────────────────────────────────────────────
@boss_bp.route('/settings', methods=['GET', 'POST'])
@boss_required
def settings():
    s = SystemSettings.query.first()
    if request.method == 'POST':
        s.student_price  = request.form['student_price']
        s.faculty_price  = request.form['faculty_price']
        s.student_both   = request.form['student_both']
        s.faculty_both   = request.form['faculty_both']
        s.one_time_price = request.form['one_time_price']
        s.lunch_start    = request.form['lunch_start']
        s.lunch_end      = request.form['lunch_end']
        s.dinner_start   = request.form['dinner_start']
        s.dinner_end     = request.form['dinner_end']
        db.session.commit()
        flash('Settings saved', 'success')
    return render_template('boss/settings.html', s=s)

# ── Helpers ───────────────────────────────────────────────────────────────────
CATEGORIES = ['Hostel', 'OneTime', 'StudentPass', 'FacultyPass', 'SpecialGuest']

def get_meal_stats_for_date(d):
    stats = {}
    for meal in ['Lunch', 'Dinner']:
        stats[meal] = {}
        total = 0
        for cat in CATEGORIES:
            r = MealCount.query.filter_by(entry_date=d, meal_type=meal, category=cat).first()
            val = r.count if r else 0
            stats[meal][cat] = val
            total += val
        stats[meal]['Total'] = total
    return stats

def get_cash_for_date(d):
    pass_cash = db.session.query(func.sum(Pass.amount_paid)).filter(
        db.cast(Pass.start_date, db.Date) == d
    ).scalar() or 0
    ot_cash = db.session.query(func.sum(OneTimeCollection.amount)).filter(
        OneTimeCollection.date == d
    ).scalar() or 0
    return float(pass_cash), float(ot_cash)

def get_rush_hours(d, meal):
    """Returns list of (hour_label, count) for a given date and meal"""
    logs = PassUsageLog.query.filter_by(date=d, meal_type=meal).all()
    buckets = {}
    for log in logs:
        if log.entry_time:
            h = log.entry_time.hour
            label = f"{h:02d}:00"
            buckets[label] = buckets.get(label, 0) + 1
    return sorted(buckets.items())

def get_low_slot_users():
    """Users with active pass and <= 5 slots remaining"""
    passes = Pass.query.filter(
        Pass.status == 'Active',
        (Pass.total_slots - Pass.used_slots) <= 5
    ).all()
    result = []
    for p in passes:
        user = User.query.get(p.user_id)
        if user:
            result.append({
                'user': user,
                'pass': p,
                'remaining': p.total_slots - p.used_slots,
                'days_left': (p.end_date - today_ist()).days
            })
    result.sort(key=lambda x: x['remaining'])
    return result

# ── Low Slots Alert (staff can also see this) ─────────────────────────────────
@boss_bp.route('/alerts')
@boss_required
def alerts():
    low_users = get_low_slot_users()
    return render_template('boss/alerts.html', low_users=low_users)

# ── Analysis: Daily ───────────────────────────────────────────────────────────
@boss_bp.route('/analysis/daily')
@boss_required
def analysis_daily():
    selected = request.args.get('date', today_ist().isoformat())
    sel_date = datetime.strptime(selected, '%Y-%m-%d').date()

    stats       = get_meal_stats_for_date(sel_date)
    pass_cash, ot_cash = get_cash_for_date(sel_date)
    grand_total = stats['Lunch']['Total'] + stats['Dinner']['Total']

    # Previous day comparison
    prev_date   = sel_date - timedelta(days=1)
    prev_stats  = get_meal_stats_for_date(prev_date)
    prev_total  = prev_stats['Lunch']['Total'] + prev_stats['Dinner']['Total']

    # Same weekday last week
    same_wday_date  = sel_date - timedelta(days=7)
    same_wday_stats = get_meal_stats_for_date(same_wday_date)
    same_wday_total = same_wday_stats['Lunch']['Total'] + same_wday_stats['Dinner']['Total']

    # Rush hours
    lunch_rush  = get_rush_hours(sel_date, 'Lunch')
    dinner_rush = get_rush_hours(sel_date, 'Dinner')

    # Holiday note
    note = DailyNote.query.filter_by(note_date=sel_date).first()

    # Low slot users alert
    low_users = get_low_slot_users()

    return render_template('boss/analysis_daily.html',
        selected=selected, sel_date=sel_date,
        stats=stats, categories=CATEGORIES,
        grand_total=grand_total,
        pass_cash=pass_cash, ot_cash=ot_cash, total_cash=pass_cash+ot_cash,
        prev_total=prev_total, prev_date=prev_date,
        same_wday_total=same_wday_total, same_wday_date=same_wday_date,
        lunch_rush=lunch_rush, dinner_rush=dinner_rush,
        note=note, low_users=low_users
    )

# ── Analysis: Monthly ─────────────────────────────────────────────────────────
@boss_bp.route('/analysis/monthly')
@boss_required
def analysis_monthly():
    month = int(request.args.get('month', today_ist().month))
    year  = int(request.args.get('year',  today_ist().year))

    # Daily rows
    rows = db.session.query(
        MealCount.entry_date,
        func.sum(db.case((MealCount.meal_type == 'Lunch',  MealCount.count), else_=0)).label('lunch'),
        func.sum(db.case((MealCount.meal_type == 'Dinner', MealCount.count), else_=0)).label('dinner'),
        func.sum(MealCount.count).label('total')
    ).filter(
        func.extract('month', MealCount.entry_date) == month,
        func.extract('year',  MealCount.entry_date) == year
    ).group_by(MealCount.entry_date).order_by(MealCount.entry_date).all()

    # Category breakdown for month
    cat_rows = db.session.query(
        MealCount.category,
        func.sum(db.case((MealCount.meal_type == 'Lunch',  MealCount.count), else_=0)).label('lunch'),
        func.sum(db.case((MealCount.meal_type == 'Dinner', MealCount.count), else_=0)).label('dinner'),
        func.sum(MealCount.count).label('total')
    ).filter(
        func.extract('month', MealCount.entry_date) == month,
        func.extract('year',  MealCount.entry_date) == year
    ).group_by(MealCount.category).all()

    # Revenue
    pass_cash = db.session.query(func.sum(Pass.amount_paid)).filter(
        func.extract('month', Pass.start_date) == month,
        func.extract('year',  Pass.start_date) == year
    ).scalar() or 0
    ot_cash = db.session.query(func.sum(OneTimeCollection.amount)).filter(
        func.extract('month', OneTimeCollection.date) == month,
        func.extract('year',  OneTimeCollection.date) == year
    ).scalar() or 0

    # Weekday averages (Mon=0 ... Sun=6)
    weekday_totals = {i: [] for i in range(7)}
    for r in rows:
        weekday_totals[r.entry_date.weekday()].append(int(r.total))
    weekday_avgs = {}
    day_names = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
    for i, name in enumerate(day_names):
        vals = weekday_totals[i]
        weekday_avgs[name] = round(sum(vals)/len(vals), 1) if vals else 0

    total_meals = sum(r.total for r in rows)
    avg_meals   = round(total_meals / len(rows), 1) if rows else 0
    busiest     = max(rows, key=lambda r: r.total) if rows else None
    slowest     = min(rows, key=lambda r: r.total) if rows else None

    # Pass count sold this month
    pass_count = Pass.query.filter(
        func.extract('month', Pass.start_date) == month,
        func.extract('year',  Pass.start_date) == year
    ).count()

    # Chart data
    chart_labels  = [r.entry_date.strftime('%d') for r in rows]
    chart_lunch   = [int(r.lunch)  for r in rows]
    chart_dinner  = [int(r.dinner) for r in rows]
    chart_wd_labels = list(weekday_avgs.keys())
    chart_wd_vals   = list(weekday_avgs.values())

    return render_template('boss/analysis_monthly.html',
        rows=rows, cat_rows=cat_rows, month=month, year=year,
        month_name=calendar.month_name[month],
        pass_cash=float(pass_cash), ot_cash=float(ot_cash),
        total_cash=float(pass_cash)+float(ot_cash),
        total_meals=total_meals, avg_meals=avg_meals,
        busiest=busiest, slowest=slowest, pass_count=pass_count,
        weekday_avgs=weekday_avgs,
        chart_labels=chart_labels, chart_lunch=chart_lunch, chart_dinner=chart_dinner,
        chart_wd_labels=chart_wd_labels, chart_wd_vals=chart_wd_vals,
        categories=CATEGORIES
    )

# ── Analysis: Range ───────────────────────────────────────────────────────────
@boss_bp.route('/analysis/range')
@boss_required
def analysis_range():
    from_str    = request.args.get('from', today_ist().replace(day=1).isoformat())
    to_str      = request.args.get('to',   today_ist().isoformat())
    meal_filter = request.args.get('meal', 'All')
    from_date   = datetime.strptime(from_str, '%Y-%m-%d').date()
    to_date     = datetime.strptime(to_str,   '%Y-%m-%d').date()

    q = db.session.query(
        MealCount.entry_date,
        func.sum(db.case((MealCount.meal_type == 'Lunch',  MealCount.count), else_=0)).label('lunch'),
        func.sum(db.case((MealCount.meal_type == 'Dinner', MealCount.count), else_=0)).label('dinner'),
        func.sum(MealCount.count).label('total')
    ).filter(MealCount.entry_date.between(from_date, to_date))
    if meal_filter != 'All':
        q = q.filter(MealCount.meal_type == meal_filter)
    rows = q.group_by(MealCount.entry_date).order_by(MealCount.entry_date.desc()).all()

    # Category breakdown
    cq = db.session.query(
        MealCount.category,
        func.sum(MealCount.count).label('total')
    ).filter(MealCount.entry_date.between(from_date, to_date))
    if meal_filter != 'All':
        cq = cq.filter(MealCount.meal_type == meal_filter)
    cat_breakdown = cq.group_by(MealCount.category).order_by(func.sum(MealCount.count).desc()).all()

    pass_cash = db.session.query(func.sum(Pass.amount_paid)).filter(
        db.cast(Pass.start_date, db.Date).between(from_date, to_date)
    ).scalar() or 0
    ot_cash = db.session.query(func.sum(OneTimeCollection.amount)).filter(
        OneTimeCollection.date.between(from_date, to_date)
    ).scalar() or 0

    total_meals = sum(r.total for r in rows)
    days_count  = len(rows)

    return render_template('boss/analysis_range.html',
        rows=rows, from_str=from_str, to_str=to_str,
        meal_filter=meal_filter, cat_breakdown=cat_breakdown,
        pass_cash=float(pass_cash), ot_cash=float(ot_cash),
        total_cash=float(pass_cash)+float(ot_cash),
        total_meals=total_meals, days_count=days_count
    )

# ── Export Excel: Monthly ─────────────────────────────────────────────────────
@boss_bp.route('/export/monthly')
@boss_required
def export_monthly():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    month = int(request.args.get('month', today_ist().month))
    year  = int(request.args.get('year',  today_ist().year))

    rows = db.session.query(
        MealCount.entry_date,
        func.sum(db.case((MealCount.meal_type == 'Lunch',  MealCount.count), else_=0)).label('lunch'),
        func.sum(db.case((MealCount.meal_type == 'Dinner', MealCount.count), else_=0)).label('dinner'),
        func.sum(MealCount.count).label('total')
    ).filter(
        func.extract('month', MealCount.entry_date) == month,
        func.extract('year',  MealCount.entry_date) == year
    ).group_by(MealCount.entry_date).order_by(MealCount.entry_date).all()

    # Category rows
    cat_rows = db.session.query(
        MealCount.category, MealCount.meal_type,
        func.sum(MealCount.count).label('total')
    ).filter(
        func.extract('month', MealCount.entry_date) == month,
        func.extract('year',  MealCount.entry_date) == year
    ).group_by(MealCount.category, MealCount.meal_type).all()

    pass_cash = float(db.session.query(func.sum(Pass.amount_paid)).filter(
        func.extract('month', Pass.start_date) == month,
        func.extract('year',  Pass.start_date) == year
    ).scalar() or 0)
    ot_cash = float(db.session.query(func.sum(OneTimeCollection.amount)).filter(
        func.extract('month', OneTimeCollection.date) == month,
        func.extract('year',  OneTimeCollection.date) == year
    ).scalar() or 0)

    wb = openpyxl.Workbook()

    # Sheet 1: Daily Meals
    ws = wb.active
    ws.title = "Daily Meals"
    header = ['Date', 'Day', 'Lunch', 'Dinner', 'Total']
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
        cell.font = Font(bold=True, color="FFFFFF")
    for r in rows:
        ws.append([
            r.entry_date.strftime('%d-%b-%Y'),
            r.entry_date.strftime('%A'),
            int(r.lunch), int(r.dinner), int(r.total)
        ])
    # Totals row
    ws.append(['', 'TOTAL',
        sum(int(r.lunch) for r in rows),
        sum(int(r.dinner) for r in rows),
        sum(int(r.total) for r in rows)
    ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    # Sheet 2: Category Breakdown
    ws2 = wb.create_sheet("Category Breakdown")
    ws2.append(['Category', 'Meal', 'Count'])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
    for r in cat_rows:
        ws2.append([r.category, r.meal_type, int(r.total)])

    # Sheet 3: Revenue
    ws3 = wb.create_sheet("Revenue")
    ws3.append(['Source', 'Amount (₹)'])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
    ws3.append(['Pass Sales', pass_cash])
    ws3.append(['One Time Collections', ot_cash])
    ws3.append(['TOTAL', pass_cash + ot_cash])
    ws3['A4'].font = Font(bold=True)
    ws3['B4'].font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    month_name = calendar.month_name[month]
    return send_file(buf, as_attachment=True,
        download_name=f'Atmiya_Mess_{month_name}_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Export Excel: Range ───────────────────────────────────────────────────────
@boss_bp.route('/export/range')
@boss_required
def export_range():
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from_str  = request.args.get('from', today_ist().replace(day=1).isoformat())
    to_str    = request.args.get('to',   today_ist().isoformat())
    from_date = datetime.strptime(from_str, '%Y-%m-%d').date()
    to_date   = datetime.strptime(to_str,   '%Y-%m-%d').date()

    rows = db.session.query(
        MealCount.entry_date,
        func.sum(db.case((MealCount.meal_type == 'Lunch',  MealCount.count), else_=0)).label('lunch'),
        func.sum(db.case((MealCount.meal_type == 'Dinner', MealCount.count), else_=0)).label('dinner'),
        func.sum(MealCount.count).label('total')
    ).filter(MealCount.entry_date.between(from_date, to_date)
    ).group_by(MealCount.entry_date).order_by(MealCount.entry_date).all()

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Meal Data"
    ws.append(['Date', 'Day', 'Lunch', 'Dinner', 'Total'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a1a2e")
    for r in rows:
        ws.append([r.entry_date.strftime('%d-%b-%Y'), r.entry_date.strftime('%A'),
                   int(r.lunch), int(r.dinner), int(r.total)])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f'Atmiya_Mess_{from_str}_to_{to_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── User History ──────────────────────────────────────────────────────────────
@boss_bp.route('/users/<int:user_id>/history')
@boss_required
def user_history(user_id):
    user   = User.query.get_or_404(user_id)
    passes = Pass.query.filter_by(user_id=user_id).order_by(Pass.id.desc()).all()
    logs   = PassUsageLog.query.filter_by(user_id=user_id).order_by(PassUsageLog.id.desc()).limit(50).all()
    return render_template('boss/user_history.html', user=user, passes=passes, logs=logs)

# ── Daily Note ────────────────────────────────────────────────────────────────
@boss_bp.route('/notes', methods=['POST'])
@boss_required
def save_note():
    note_date  = datetime.strptime(request.form['note_date'], '%Y-%m-%d').date()
    label      = request.form.get('label', '')
    is_holiday = request.form.get('is_holiday') == 'on'
    note = DailyNote.query.filter_by(note_date=note_date).first()
    if note:
        note.label = label; note.is_holiday = is_holiday
    else:
        note = DailyNote(note_date=note_date, label=label, is_holiday=is_holiday)
        db.session.add(note)
    db.session.commit()
    flash('Note saved', 'success')
    return redirect(request.referrer or url_for('boss.analysis_daily'))